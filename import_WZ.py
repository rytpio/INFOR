import pandas as pd
from dicts_n_lists import file_import_map
from openpyxl import load_workbook
import os.path
import os
import warnings
import logging
from sys import stdout


def logfile_lastline(logfile_txt):
    with open(logfile_txt, 'rb') as f:
        f.seek(os.SEEK_END-2, os.SEEK_END-1)  # nie działa
        while f.read(1) != b'\n':
            f.seek(os.SEEK_END-2, os.SEEK_END-1)
        print(f.readline().decode())


def import_struktur(project: str, standard: str, path_read: str, path_save: str) -> pd.DataFrame():
    # warnings.simplefilter("ignore")
    #
    # formatter = logging.Formatter(fmt='%(asctime)s %(message)s', datefmt='%d-%m-%Y %H:%M:%S')
    # handler = logging.FileHandler('AVZ/log.txt', mode='w')
    # handler.setFormatter(formatter)
    # screen_handler = logging.StreamHandler(stream=stdout)
    # screen_handler.setFormatter(formatter)
    # logger = logging.getLogger('wz')
    # logger.setLevel(logging.DEBUG)
    # logger.addHandler(handler)
    #
    # logger.info("kisters links started")  # calculate time n rows x time registered
    # logfile_lastline('log.txt')

    # STAPS
    wb = load_workbook(path_read, data_only=True, keep_links=True)
    ws = wb.active
    link_list = []
    col = list(file_import_map.wz_col_dict_STAPS.keys()).index('Kistersviewerlink') - 1
    for row in range(6, ws.max_row + 1, 1):
        try:
            link_list.append(ws.cell(column=col, row=row).hyperlink.target)
        except:
            link_list.append(None)
    wb.close()

    #  logger.info('kisters finished \n  wz upload start')
    df_strukt = pd.read_excel(path_read, dtype='string', header=5)  # dodac pierwszy arkusz
    match standard:
        case "STAG":
            df_strukt.rename(columns=file_import_map.wz_col_dict_STAG, inplace=True)
        case "STAPS":
            df_strukt.rename(columns=file_import_map.wz_col_dict_STAPS, inplace=True)
            df_strukt = df_strukt[file_import_map.wz_col_short_list_STAPS]
        case _:
            pass
            #  logger.error("Brak standardu określonego.")
            # TODO: log, catch missing and end
    #  logger.info('wz upload finished \n level 123 description')

    #  lvl(1, 2, 3)
    lvl_1 = []
    lvl_2 = []
    lvl_3 = []
    for row in df_strukt.itertuples():
        if int(row.level) == 1:
            lvl1 = row.description_1
            lvl2 = lvl3 = None
        if int(row.level) == 2:
            lvl2 = row.description_1
            lvl3 = None
        if int(row.level) == 3:
            lvl3 = row.description_1
        lvl_1.append(lvl1)
        lvl_2.append(lvl2)
        lvl_3.append(lvl3)

    df_strukt['viewer_link'] = pd.Series(link_list)
    df_strukt['lvl1'] = pd.Series(lvl_1)
    df_strukt['lvl2'] = pd.Series(lvl_2)
    df_strukt['lvl3'] = pd.Series(lvl_3)
    df_strukt['project'] = project

    def get_car(x):  # zrobione pod strukture BLS; zrobic abstrakcje przed uzyciem do innych projektow

        if 'Wagon ' in x:
            result = x[6:7]  # zalozenie ze wagon jest 1-literowy, nazwy są stale;
            # naprawde to trzeba slownika i flagi co do wyciagania wagonow
        elif 'Wózki,' in x:
            result = 'DG'
        elif 'IBS' in x:
            result = 'x'  # 'IBS'
        elif 'allgemein' in x:
            result = 'x'  # 'general'
        else:
            result = 'x'  # 'root'

        return result

    #  logger.info('level described \n car description')
    df_strukt['car'] = df_strukt['wz_header'].apply(lambda x: get_car(str(x)))
    #  df_strukt = df_strukt[df_strukt['car'] != 'x']  # odsiew zbednych
    df_strukt['car_stadler_id'] = df_strukt.apply(lambda row_x: f'{row_x.car}_{row_x.stadler_id}', axis=1)
    # df_strukt['car_stadler_mnrpath_id'] = df_strukt.apply(lambda row: f'{row.car}_{row.stadler_id}_{row.mnr_path}',
    #                                                       axis=1)

    #  logger.info('finished & saving')
    df_strukt.to_excel(path_save, index=False)

    return df_strukt


